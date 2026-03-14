#!/usr/bin/env python
"""
安装验证和启动检查脚本。

用法：
    python setup_check.py              # 完整检查
    python setup_check.py --quick      # 快速检查（仅依赖）
    python setup_check.py --verbose    # 详细输出
"""

import os
import sys
from pathlib import Path


def print_header(text: str):
    print("\n" + "=" * 60)
    print(text)
    print("=" * 60)


def print_check(name: str, passed: bool, details: str = ""):
    status = "✅" if passed else "❌"
    print(f"{status} {name}")
    if details and not passed:
        print(f"   → {details}")


def check_python_version() -> bool:
    """检查 Python 版本 >= 3.8"""
    major, minor = sys.version_info[:2]
    passed = (major == 3 and minor >= 8) or (major > 3)
    print_check(
        f"Python version (required: 3.8+, current: {major}.{minor})",
        passed,
        f"Current version: {major}.{minor}" if not passed else ""
    )
    return passed


def check_dependencies(verbose: bool = False) -> tuple[bool, list]:
    """检查必需依赖"""
    core_packages = [
        ("pandas", "pandas"),
        ("numpy", "numpy"),
        ("sklearn", "scikit-learn"),
        ("openpyxl", "openpyxl"),
        ("xlsxwriter", "xlsxwriter"),
        ("joblib", "joblib"),
    ]
    
    optional_packages = [
        ("xgboost", "xgboost"),
    ]
    
    missing_required = []
    missing_optional = []
    installed = []
    
    for import_name, pkg_name in core_packages:
        try:
            __import__(import_name)
            installed.append(pkg_name)
            if verbose:
                print_check(f"  {pkg_name}", True)
        except ImportError:
            missing_required.append(pkg_name)
            print_check(f"  {pkg_name}", False)
    
    if verbose:
        print("\n  Optional packages:")
        for import_name, pkg_name in optional_packages:
            try:
                __import__(import_name)
                installed.append(pkg_name)
                print_check(f"  {pkg_name} (optional)", True)
            except ImportError:
                missing_optional.append(pkg_name)
                print_check(f"  {pkg_name} (optional)", False, "Not installed (optional)")
    
    all_passed = len(missing_required) == 0
    return all_passed, missing_required, missing_optional


def check_data_file(verbose: bool = False) -> bool:
    """检查示例数据文件"""
    script_dir = Path(__file__).resolve().parent
    data_path = script_dir / "data" / "600698_5Y_daily_qfq_20210302_20260302.xlsx"
    
    exists = data_path.exists()
    print_check(f"Sample data file", exists, str(data_path) if not exists else "")
    
    if exists and verbose:
        import openpyxl
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(data_path), read_only=True)
            sheet = wb.sheetnames[0]
            print(f"   → Sheet: {sheet}")
            wb.close()
        except Exception as e:
            print(f"   → Warning: Cannot read Excel file: {e}")
    
    return exists


def check_output_dir(verbose: bool = False) -> bool:
    """检查输出目录"""
    output_dir = Path(__file__).resolve().parent / "output"
    
    try:
        output_dir.mkdir(exist_ok=True)
        test_file = output_dir / ".write_test"
        test_file.write_text("test")
        test_file.unlink()
        print_check(f"Output directory (writable)", True, str(output_dir))
        return True
    except Exception as e:
        print_check(f"Output directory", False, str(e))
        return False


def check_project_structure(verbose: bool = False) -> bool:
    """检查项目结构"""
    required_files = [
        "main_cli.py",
        "data_loader.py",
        "feature_dpoint.py",
        "model_builder.py",
        "backtester_engine.py",
        "search_engine.py",
        "trainer_optimizer.py",
        "metrics.py",
        "splitter.py",
        "reporter.py",
        "persistence.py",
        "constants.py",
    ]
    
    all_exist = True
    for file in required_files:
        exists = (Path(__file__).resolve().parent / file).exists()
        if verbose:
            print_check(f"  {file}", exists)
        if not exists:
            all_exist = False
    
    return all_exist


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description="Installation and setup verification")
    parser.add_argument("--quick", action="store_true", help="Quick check (dependencies only)")
    parser.add_argument("--verbose", "-v", action="store_true", help="Verbose output")
    args = parser.parse_args()
    
    print_header("A-Share Dpoint Trader 2.0 - Setup Verification")
    
    results = []
    
    # 1. Python version
    print("\n[1/5] Checking Python version...")
    results.append(("Python version", check_python_version()))
    
    # 2. Dependencies
    print("\n[2/5] Checking dependencies...")
    deps_passed, missing_req, missing_opt = check_dependencies(args.verbose)
    results.append(("Dependencies", deps_passed))
    if missing_req:
        print(f"\n  💡 Install missing: pip install {' '.join(missing_req)}")
    if missing_opt:
        print(f"\n  ℹ️  Optional (not installed): {' '.join(missing_opt)}")
    
    if args.quick:
        # Quick mode: only check Python and dependencies
        print_header("Summary (Quick Check)")
    else:
        # 3. Data file
        print("\n[3/5] Checking sample data file...")
        results.append(("Sample data", check_data_file(args.verbose)))
        
        # 4. Output directory
        print("\n[4/5] Checking output directory...")
        results.append(("Output directory", check_output_dir(args.verbose)))
        
        # 5. Project structure
        print("\n[5/5] Checking project structure...")
        results.append(("Project structure", check_project_structure(args.verbose)))
        
        print_header("Summary")
    
    # Summary
    total = len(results)
    passed = sum(1 for _, p in results if p)
    
    for name, result in results:
        print_check(name, result)
    
    print(f"\nResult: {passed}/{total} checks passed")
    
    if passed == total:
        print("\n✅ All checks passed! You're ready to run.")
        print("\nNext step:")
        print("  python main_cli.py --runs 10  # Quick test with 10 iterations")
        return 0
    else:
        print("\n❌ Some checks failed. Please fix the issues above.")
        print("\nTo install missing dependencies:")
        print("  pip install -r requirements.txt")
        return 1


if __name__ == "__main__":
    sys.exit(main())
